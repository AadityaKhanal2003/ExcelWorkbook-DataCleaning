#Gaurab Baral, Aditya Khanal Group 8, Lab 4. 
#Importing essential libraries
import csv
import openpyxl
import xlrd
#get the required excel workbook
path = "Lab4Data.xlsx"
wb = openpyxl.load_workbook(path, read_only=False)
ws = wb.worksheets[1]


#range of essential columns
countries_range = ws["B15:B211"]
clabour_range = ws["E15:E211"]
clabourM_range = ws["G15:G211"]
clabourW_range = ws["I15:I211"]
cmarraigeb15_range= ws["K15:K211"]
cmarraigeb18_range= ws["M15:M211"]
bregistration_range = ws["O15:O211"]
total_fgm__rangeW= ws["Q15:Q211"]
total_fgm__rangeG= ws["S15:S211"]
total_fgm__rangeSGM= ws["U15:U211"]
total_jow_rangeM= ws["W15:W211"]
total_jow_rangeF= ws["Y15:Y211"]
vdiscpline_rangeM = ws["AC15:AC211"]
vdiscpline_range = ws["AA15:AA211"]
vdiscpline_rangeF = ws["AE15:AE211"]

#all the values of the 14 columns, these all include 197 values.
country_values = [cell[0].value for cell in countries_range] #197 countries
clabour_values = [cell[0].value for cell in clabour_range]#197 clabour_total
clabourM_values = [cell[0].value for cell in clabourM_range]#197 c_labour_male
clabourW_values = [cell[0].value for cell in clabourW_range] #197 c_labout_female
cmarraigeb15_values = [cell[0].value for cell in cmarraigeb15_range]#197 c_marriage before 15
cmarraigeb18_values = [cell[0].value for cell in cmarraigeb18_range]#197 c_marriage before 18
bregistration_values = [cell[0].value for cell in bregistration_range] #197 total birth registration
total_fgmW_values = [cell[0].value for cell in total_fgm__rangeW] #197 total FGM prevelance in women
total_fgmG_values = [cell[0].value for cell in total_fgm__rangeG] #197 total FGM prevelance in women
total_fgmSGM_values = [cell[0].value for cell in total_fgm__rangeSGM]#197 total FGM prevelance in support
total_jowM_values = [cell[0].value for cell in total_jow_rangeM] #197 justify wife beating female
total_jowF_values = [cell[0].value for cell in total_jow_rangeF] #197 justify wife beating male
vdiscpline_valuesM = [cell[0].value for cell in vdiscpline_rangeM]#197 Violent Disclipine male
vdiscpline_valuesF = [cell[0].value for cell in vdiscpline_rangeF]#197 Violent Disclipine female
vdiscpline_values = [cell[0].value for cell in vdiscpline_range]#197 Violent Disclipine total

#I created a list that included all the columns so that it will be easier for iteration
catagories_list = [clabour_values,clabourM_values,clabourW_values,cmarraigeb15_values,cmarraigeb18_values,bregistration_values,
                   total_fgmW_values,total_fgmG_values,total_fgmSGM_values,total_jowM_values,total_jowF_values,vdiscpline_values,
                   vdiscpline_valuesM,vdiscpline_valuesF]
#title of each column
title = ['Child_labour_total', 'Child_labour_male', 'Child_labour_female', 'Children married by 15', 
         'Children married by 18', 'Birth registration total', 'FGM prevelance in women', 'FGM prevalance in girls','FGM prevelance in support groups', 
         'Justify wife beating male', 'Justify wife beating female', 'Violent discipline total', 'Violent discipline male', 
         'Violent discipline female']
#title for csv file columns
header1 = ["CountryName", "CategoryName", "CategoryTotal"]
#dict for csv file
data_dict = {
    header1[0] : [],
    header1[1] :[],
    header1[2] :[]
}
#count to count the number of rows
count = 0
rows = []
for i in range(len(country_values)):
    for j in range(14):  # As there are 14 categories
        if "–" in str(catagories_list[j][i]) or catagories_list[j][i] == 0: #Remove all 0 and all "-
            continue
        row = {
            header1[0]: country_values[i],          #Country Name
            header1[1]: title[j],                   #Category Name
            header1[2]: catagories_list[j][i]       #Category Total
        }
        count+=1
        rows.append(row)
#Writing it in a csv file
output_file = "8_Lab4.csv"
with open(output_file, "w", newline="") as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=header1) #write title as header1
    writer.writeheader()
    for row in rows:
        writer.writerow(row)                              #add each row in a new row in csv file

print("The total number of rows in the csv file is ",count)

#Output:
#The total number of rows in the csv file is  1082